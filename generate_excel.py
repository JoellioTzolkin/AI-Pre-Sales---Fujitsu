#!/usr/bin/env python3
"""
Generate an Excel workbook (layout similar to "Osvaldo") from an estimate.json created by Claude.

Usage:
  python generate_excel.py --input samples/example_estimate.json --output out.xlsx --start-date 2026-01-15

Notes:
- This script intentionally prioritizes consistent layout + auditability over “perfect scheduling”.
- If you want real dates/dependencies, extend the JSON to carry task start/finish or parse MS Project XML directly.
"""
from __future__ import annotations
import argparse, json, math, datetime
from collections import defaultdict
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

THIN = Side(style="thin", color="A0A0A0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def parse_args():
    p = argparse.ArgumentParser()
    p.add_argument("--input", required=True, help="estimate.json")
    p.add_argument("--output", required=True, help="Output .xlsx")
    p.add_argument("--start-date", default=None, help="YYYY-MM-DD (optional, for a simple phase timeline)")
    p.add_argument("--hours-per-day", type=float, default=None, help="Override hours_per_day from JSON")
    p.add_argument("--days-per-week", type=int, default=5)
    p.add_argument("--default-phase-fte", type=float, default=2.0, help="Used only for simple timeline (sequential phases)")
    return p.parse_args()

def safe_date(s: str) -> datetime.date:
    return datetime.date.fromisoformat(s)

def autosize(ws, min_col=1, max_col=50, extra=2):
    for col in range(min_col, max_col+1):
        max_len = 0
        col_letter = get_column_letter(col)
        for cell in ws[col_letter]:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        if max_len:
            ws.column_dimensions[col_letter].width = min(60, max_len + extra)

def make_table(ws, ref: str, name: str):
    tab = Table(displayName=name, ref=ref)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

def main():
    args = parse_args()
    with open(args.input, "r", encoding="utf-8") as f:
        est = json.load(f)

    hours_per_day = float(args.hours_per_day or est.get("hours_per_day", 8))

    wb = Workbook()
    wb.remove(wb.active)

    # -------------------------
    # Sheet: Parameters
    # -------------------------
    ws = wb.create_sheet("Parameters")
    ws["A1"] = "roadmap_name"; ws["B1"] = est.get("roadmap_name","")
    ws["A2"] = "estimation_mode"; ws["B2"] = est.get("estimation_mode","")
    ws["A3"] = "hours_per_day"; ws["B3"] = hours_per_day
    ws["A4"] = "days_per_week"; ws["B4"] = args.days_per_week
    ws["A5"] = "default_phase_fte"; ws["B5"] = args.default_phase_fte
    ws["A6"] = "start_date"; ws["B6"] = args.start_date or ""
    for r in range(1,7):
        ws[f"A{r}"].font = Font(bold=True)
    autosize(ws,1,2)

    # -------------------------
    # Sheet: Effort by Phase/Role
    # -------------------------
    ws = wb.create_sheet("Effort_PhaseRole")
    ws.freeze_panes = "B2"
    ws["A1"] = "Phase"
    # collect roles
    roles = sorted({role for phase, m in est.get("effort_matrix_hours", {}).items() for role in m.keys()})
    for j, role in enumerate(roles, start=2):
        ws.cell(row=1, column=j, value=role)
    ws.cell(row=1, column=2+len(roles), value="Total Hours")
    ws.cell(row=1, column=3+len(roles), value="Total Days")

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for c in range(1, 4+len(roles)):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    phases = list(est.get("effort_matrix_hours", {}).keys())
    for i, phase_name in enumerate(phases, start=2):
        ws.cell(row=i, column=1, value=phase_name).border = BORDER
        total = 0.0
        for j, role in enumerate(roles, start=2):
            v = float(est["effort_matrix_hours"].get(phase_name, {}).get(role, 0.0))
            total += v
            c = ws.cell(row=i, column=j, value=v)
            c.number_format = "0.0"
            c.border = BORDER
        ws.cell(row=i, column=2+len(roles), value=total).number_format = "0.0"
        ws.cell(row=i, column=2+len(roles)).border = BORDER
        ws.cell(row=i, column=3+len(roles), value=total / hours_per_day).number_format = "0.0"
        ws.cell(row=i, column=3+len(roles)).border = BORDER

    autosize(ws, 1, 4+len(roles))
    make_table(ws, f"A1:{get_column_letter(3+len(roles))}{len(phases)+1}", "TblPhaseRole")

    # -------------------------
    # Sheet: Deliverables
    # -------------------------
    ws = wb.create_sheet("Deliverables")
    ws.freeze_panes = "A2"
    headers = ["accelerator_id","title","phases","estimated_hours","estimated_days","source_url","accelerator_type","access_level"]
    for j,h in enumerate(headers, start=1):
        ws.cell(row=1, column=j, value=h)
        cell = ws.cell(row=1, column=j)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    delivs = est.get("deliverables", [])
    delivs_sorted = sorted(delivs, key=lambda d: float(d.get("estimated_hours",0)), reverse=True)
    for i, d in enumerate(delivs_sorted, start=2):
        ws.cell(row=i, column=1, value=d.get("accelerator_id","")).border = BORDER
        ws.cell(row=i, column=2, value=d.get("title","")).border = BORDER
        ws.cell(row=i, column=3, value=", ".join(d.get("phases",[]))).border = BORDER
        hrs = float(d.get("estimated_hours",0))
        ws.cell(row=i, column=4, value=hrs).number_format="0.0"
        ws.cell(row=i, column=4).border = BORDER
        ws.cell(row=i, column=5, value=hrs / hours_per_day).number_format="0.0"
        ws.cell(row=i, column=5).border = BORDER
        ws.cell(row=i, column=6, value=d.get("source_url","")).border = BORDER
        ws.cell(row=i, column=7, value=d.get("accelerator_type","")).border = BORDER
        ws.cell(row=i, column=8, value=d.get("access_level","")).border = BORDER

    autosize(ws,1,8)
    make_table(ws, f"A1:H{len(delivs_sorted)+1}", "TblDeliverables")

    # -------------------------
    # Sheet: Tasks (optional but helps audit)
    # -------------------------
    ws = wb.create_sheet("Tasks")
    ws.freeze_panes = "A2"
    headers = ["phase","sequence","title","leading_workstream","roles","accelerators","estimated_hours_total"]
    for j,h in enumerate(headers, start=1):
        ws.cell(row=1, column=j, value=h)
        cell = ws.cell(row=1, column=j)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    row = 2
    for ph in est.get("phases", []):
        phase_name = ph.get("phase_name") or ph.get("phase_number","")
        for t in ph.get("tasks", []):
            ws.cell(row=row, column=1, value=phase_name).border=BORDER
            ws.cell(row=row, column=2, value=t.get("sequence","")).border=BORDER
            ws.cell(row=row, column=3, value=t.get("title","")).border=BORDER
            ws.cell(row=row, column=4, value=t.get("leading_workstream","")).border=BORDER
            ws.cell(row=row, column=5, value=", ".join(t.get("roles",[]))).border=BORDER
            ws.cell(row=row, column=6, value=", ".join(t.get("accelerators",[]))).border=BORDER
            ws.cell(row=row, column=7, value=float(t.get("estimated_hours_total",0))).number_format="0.0"
            ws.cell(row=row, column=7).border=BORDER
            row += 1

    autosize(ws,1,7)
    make_table(ws, f"A1:G{row-1}", "TblTasks")

    # -------------------------
    # Sheet: Effort_Estimation (simple timeline, optional)
    # -------------------------
    ws = wb.create_sheet("Effort_Estimation")
    ws.freeze_panes = "C4"

    # Title
    ws["A1"] = "Effort Estimation (POC)"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"] = "Roadmap:"
    ws["B2"] = est.get("roadmap_name","")
    ws["A3"] = "Assumption:"
    ws["B3"] = "Timeline below is a simple sequential phase view (not dependency-accurate)."

    # timeline only if start-date provided
    start_date = safe_date(args.start_date) if args.start_date else None

    # header rows
    ws["A4"] = "Phase"; ws["B4"] = "Total Hours"
    ws["A4"].font = Font(bold=True); ws["B4"].font = Font(bold=True)

    # compute phase durations
    phase_rows = []
    for ph in est.get("phases", []):
        name = ph.get("phase_name") or ph.get("phase_number","")
        total_hours = float(ph.get("total_hours",0))
        duration_days = math.ceil(total_hours / (hours_per_day * args.default_phase_fte)) if args.default_phase_fte>0 else 0
        phase_rows.append((name, total_hours, duration_days))

    # Determine total weeks
    days_per_week = args.days_per_week
    total_days = sum(d for _,_,d in phase_rows)
    total_weeks = max(1, math.ceil(total_days / days_per_week)) if start_date else 0

    # Write week headers
    if start_date:
        for w in range(total_weeks):
            col = 3 + w
            ws.cell(row=4, column=col, value=f"W{w+1}")
            ws.cell(row=4, column=col).font = Font(bold=True)
            ws.cell(row=4, column=col).alignment = Alignment(horizontal="center")
            ws.cell(row=4, column=col).border = BORDER
            ws.column_dimensions[get_column_letter(col)].width = 4

    # Fill rows
    fills = [
        PatternFill("solid", fgColor="BDD7EE"),
        PatternFill("solid", fgColor="C6E0B4"),
        PatternFill("solid", fgColor="F8CBAD"),
        PatternFill("solid", fgColor="D9D2E9"),
        PatternFill("solid", fgColor="FFE699"),
    ]

    current_day = 0
    for i,(name, hrs, dur_days) in enumerate(phase_rows, start=5):
        ws.cell(row=i, column=1, value=name).border=BORDER
        ws.cell(row=i, column=2, value=hrs).number_format="0.0"
        ws.cell(row=i, column=2).border=BORDER
        if start_date:
            start_w = current_day // days_per_week
            end_day = current_day + max(0, dur_days)
            end_w = max(start_w, math.ceil(end_day / days_per_week) - 1)
            fill = fills[(i-5) % len(fills)]
            for w in range(start_w, end_w+1):
                col = 3 + w
                c = ws.cell(row=i, column=col, value="")
                c.fill = fill
                c.border = BORDER
            current_day = end_day

    # Borders for non-filled timeline cells
    if start_date:
        for r in range(5, 5+len(phase_rows)):
            for col in range(3, 3+total_weeks):
                c = ws.cell(row=r, column=col)
                if c.border is None:
                    c.border = BORDER

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 12

    wb.save(args.output)

if __name__ == "__main__":
    main()
